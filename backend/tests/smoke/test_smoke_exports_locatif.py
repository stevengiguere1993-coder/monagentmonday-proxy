"""Smoke — exports du pôle Gestion locative (2026-09-09).

- tableaux CSV (BOM + « ; » + en-têtes) et XLSX (ouvrable par openpyxl,
  en-têtes en gras) sur /immobilier/exports/* — les lignes sont celles
  de la page Paiements (interne + gestion externe) ;
- zip des documents d'un locataire : index.csv à la racine, fichiers
  nommés AAAA-MM-JJ_type_titre_id.pdf, sous-dossier par bail,
  catégorie « dossier » qui exclut les simples communications ;
- garde-fous 413 (MAX_DOCS / MAX_ZIP_BYTES, monkeypatchés).
"""
from __future__ import annotations

import csv
import io
import zipfile
from datetime import date, datetime, timedelta, timezone

import pytest

from app.api.v1.endpoints import immobilier_exports
from app.models.immobilier import (
    Bail,
    BailStatus,
    ImmDocument,
    Immeuble,
    Locataire,
    Logement,
    LogementStatus,
    PaiementLoyer,
)

from .conftest import TestSessionLocal

_PDF = b"%PDF-1.4 smoke export document"
_BOM = "\ufeff"


def _mois_courant() -> str:
    return datetime.now(timezone.utc).date().strftime("%Y-%m")


@pytest.fixture(scope="module")
def seed(run, seeded_users) -> dict:
    """Un immeuble INTERNE (bail actif + paiement du mois + 3 documents)
    et un immeuble en GESTION EXTERNE (logement occupé, rien reçu)."""

    async def _seed() -> dict:
        async with TestSessionLocal() as s:
            imm = Immeuble(
                name="Immeuble Smoke Export",
                address="12 rue de l'Export",
                is_active=True,
            )
            ext = Immeuble(
                name="Tour Export Externe",
                address="34 rue Déléguée",
                is_active=True,
                gestion_externe=True,
            )
            s.add_all([imm, ext])
            await s.flush()
            lg = Logement(
                immeuble_id=imm.id,
                numero="EXP-1",
                status=LogementStatus.OCCUPE.value,
            )
            lg_ext = Logement(
                immeuble_id=ext.id,
                numero="EXT-9",
                status=LogementStatus.OCCUPE.value,
                loyer_demande=750.0,
            )
            loc = Locataire(
                full_name="Éliane Exporté",
                email="eliane@example.com",
                phone="514 555-0199",
            )
            s.add_all([lg, lg_ext, loc])
            await s.flush()
            mois = datetime.now(timezone.utc).date().replace(day=1)
            debut = (mois - timedelta(days=1)).replace(day=1)
            debut = (debut - timedelta(days=1)).replace(day=1)
            bail = Bail(
                logement_id=lg.id,
                locataire_id=loc.id,
                date_debut=debut,
                date_fin=debut + timedelta(days=730),
                loyer_mensuel=1200.0,
                status=BailStatus.ACTIF.value,
            )
            s.add(bail)
            await s.flush()
            s.add(
                PaiementLoyer(
                    bail_id=bail.id,
                    mois_couvert=mois,
                    montant=1200.0,
                    paye_le=mois,
                    created_at=datetime.now(timezone.utc),
                )
            )
            docs = [
                # Pièce du dossier, rattachée au bail (sous-dossier).
                ImmDocument(
                    bail_id=bail.id,
                    locataire_id=loc.id,
                    immeuble_id=imm.id,
                    type="bail",
                    titre="Bail signé 2026",
                    source="importe",
                    filename="bail.pdf",
                    pdf_blob=_PDF,
                    signed_at=datetime.now(timezone.utc),
                    signed_by_name="Éliane Exporté",
                ),
                # Pièce importée directement sur le locataire (racine).
                ImmDocument(
                    locataire_id=loc.id,
                    type="autre",
                    titre="Preuve d'assurance : 2026/2027",
                    source="importe",
                    filename="assurance.pdf",
                    pdf_blob=_PDF,
                ),
                # Simple communication générée (hors « dossier »).
                ImmDocument(
                    bail_id=bail.id,
                    locataire_id=loc.id,
                    immeuble_id=imm.id,
                    type="rappel_paiement",
                    titre="Rappel de paiement",
                    source="genere",
                    pdf_blob=_PDF,
                    envoye_le=datetime.now(timezone.utc),
                    envoye_a="eliane@example.com",
                ),
            ]
            s.add_all(docs)
            await s.commit()
            return {
                "immeuble_id": imm.id,
                "immeuble_ext_id": ext.id,
                "logement_id": lg.id,
                "locataire_id": loc.id,
                "bail_id": bail.id,
                "mois": mois.strftime("%Y-%m"),
            }

    return run(_seed())


def _lignes_csv(content: bytes) -> list[list[str]]:
    texte = content.decode("utf-8")
    assert texte.startswith(_BOM), "BOM UTF-8 attendu en tête du CSV"
    return list(csv.reader(io.StringIO(texte[len(_BOM):]), delimiter=";"))


# ── Tableaux ───────────────────────────────────────────────────────────


def test_export_paiements_csv(client, auth_headers, seed):
    r = client.get(
        f"/api/v1/immobilier/exports/paiements?mois={seed['mois']}&fmt=csv",
        headers=auth_headers,
    )
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("text/csv")
    cd = r.headers["content-disposition"]
    assert cd.startswith('attachment; filename="kratos_paiements_')
    assert cd.endswith('.csv"')

    lignes = _lignes_csv(r.content)
    assert lignes[0][:4] == ["Mois", "Immeuble", "Logement", "Locataire"]
    assert "Gestion externe" in lignes[0]
    assert lignes[0][-1] == "Nb relances"
    idx = {h: i for i, h in enumerate(lignes[0])}

    # Ligne interne : le bail payé du mois, avec ses coordonnées.
    interne = next(
        l for l in lignes[1:] if l[idx["Locataire"]] == "Éliane Exporté"
    )
    assert interne[idx["Mois"]] == seed["mois"]
    assert interne[idx["Immeuble"]] == "Immeuble Smoke Export"
    assert interne[idx["Logement"]] == "EXP-1"
    assert interne[idx["Courriel"]] == "eliane@example.com"
    assert interne[idx["Gestion externe"]] == "non"
    assert interne[idx["Loyer attendu"]] == "1200.00"
    assert interne[idx["Payé"]] == "1200.00"
    assert interne[idx["Solde du mois"]] == "0.00"
    assert interne[idx["État"]] in ("Payé", "Partiel")

    # Ligne de gestion externe : logement occupé, rien reçu.
    externe = next(
        l for l in lignes[1:] if l[idx["Logement"]] == "EXT-9"
    )
    assert externe[idx["Immeuble"]] == "Tour Export Externe"
    assert externe[idx["Gestion externe"]] == "oui"
    assert externe[idx["Loyer attendu"]] == "750.00"
    assert externe[idx["État"]] in ("En retard", "En attente")


def test_export_paiements_filtre_immeuble_et_periode(
    client, auth_headers, seed
):
    # Filtre immeuble : seul l'interne reste.
    r = client.get(
        "/api/v1/immobilier/exports/paiements"
        f"?mois={seed['mois']}&immeuble_id={seed['immeuble_id']}",
        headers=auth_headers,
    )
    assert r.status_code == 200, r.text
    lignes = _lignes_csv(r.content)
    assert all(l[1] == "Immeuble Smoke Export" for l in lignes[1:])
    assert len(lignes) >= 2

    # Période du/au : deux mois → la ligne du bail apparaît deux fois.
    mois = date.fromisoformat(seed["mois"] + "-01")
    prec = (mois - timedelta(days=1)).replace(day=1).strftime("%Y-%m")
    r = client.get(
        "/api/v1/immobilier/exports/paiements"
        f"?du={prec}&au={seed['mois']}&immeuble_id={seed['immeuble_id']}",
        headers=auth_headers,
    )
    assert r.status_code == 200, r.text
    lignes = _lignes_csv(r.content)
    mois_vus = sorted({l[0] for l in lignes[1:]})
    assert mois_vus == [prec, seed["mois"]]
    assert f"kratos_paiements_{prec}_au_{seed['mois']}_" in (
        r.headers["content-disposition"]
    )

    # Période mal formée → 400 explicite.
    assert (
        client.get(
            "/api/v1/immobilier/exports/paiements?du=2026-01",
            headers=auth_headers,
        ).status_code
        == 400
    )
    assert (
        client.get(
            "/api/v1/immobilier/exports/paiements?mois=2026-13",
            headers=auth_headers,
        ).status_code
        == 400
    )
    # Format inconnu → 400.
    assert (
        client.get(
            f"/api/v1/immobilier/exports/paiements?mois={seed['mois']}&fmt=pdf",
            headers=auth_headers,
        ).status_code
        == 400
    )


def test_export_paiements_xlsx(client, auth_headers, seed):
    from openpyxl import load_workbook

    r = client.get(
        f"/api/v1/immobilier/exports/paiements?mois={seed['mois']}&fmt=xlsx",
        headers=auth_headers,
    )
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert r.headers["content-disposition"].endswith('.xlsx"')
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    lignes = list(ws.iter_rows(values_only=False))
    entetes = [c.value for c in lignes[0]]
    assert entetes[:3] == ["Mois", "Immeuble", "Logement"]
    assert all(c.font.bold for c in lignes[0])
    idx = {h: i for i, h in enumerate(entetes)}
    interne = next(
        l for l in lignes[1:]
        if l[idx["Locataire"]].value == "Éliane Exporté"
    )
    # Les montants restent NUMÉRIQUES dans Excel.
    assert interne[idx["Loyer attendu"]].value == 1200.0
    assert interne[idx["Payé"]].value == 1200.0
    assert interne[idx["Gestion externe"]].value == "non"


@pytest.mark.parametrize(
    "sujet,premiere_colonne",
    [
        ("locataires", "ID"),
        ("baux", "Immeuble"),
        ("logements", "Immeuble"),
        ("immeubles", "ID"),
        ("depots", "Immeuble"),
    ],
)
def test_exports_tableaux_csv_et_xlsx(
    client, auth_headers, seed, sujet, premiere_colonne
):
    from openpyxl import load_workbook

    r = client.get(
        f"/api/v1/immobilier/exports/{sujet}"
        f"?immeuble_id={seed['immeuble_id']}&fmt=csv",
        headers=auth_headers,
    )
    assert r.status_code == 200, r.text
    lignes = _lignes_csv(r.content)
    assert lignes[0][0] == premiere_colonne
    assert len(lignes) >= 2, f"aucune ligne pour {sujet}"
    assert f'filename="kratos_{sujet}_' in r.headers["content-disposition"]
    # La donnée de l'immeuble seed est bien là.
    corps = "\n".join(";".join(l) for l in lignes[1:])
    assert (
        "Immeuble Smoke Export" in corps
        or "Éliane Exporté" in corps
        or "EXP-1" in corps
    )

    x = client.get(
        f"/api/v1/immobilier/exports/{sujet}"
        f"?immeuble_id={seed['immeuble_id']}&fmt=xlsx",
        headers=auth_headers,
    )
    assert x.status_code == 200, x.text
    ws = load_workbook(io.BytesIO(x.content)).active
    assert ws.cell(row=1, column=1).value == premiere_colonne
    assert ws.max_row >= 2


def test_exports_refuses_sans_volet(client, employee_headers, seed):
    """Même contrôle d'accès que les endpoints voisins : un employé sans
    volet immobilier est refusé."""
    r = client.get(
        "/api/v1/immobilier/exports/locataires", headers=employee_headers
    )
    assert r.status_code in (401, 403)


# ── Zip de documents ───────────────────────────────────────────────────


def test_zip_locataire_index_et_fichiers(client, auth_headers, seed):
    r = client.get(
        f"/api/v1/immobilier/locataires/{seed['locataire_id']}"
        "/documents.zip?categorie=tout",
        headers=auth_headers,
    )
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == "application/zip"
    assert r.headers["content-disposition"].startswith(
        'attachment; filename="kratos_documents_locataire_'
    )
    zf = zipfile.ZipFile(io.BytesIO(r.content))
    noms = zf.namelist()
    assert "index.csv" in noms
    fichiers = [n for n in noms if n != "index.csv"]
    assert len(fichiers) == 3

    # Nommage AAAA-MM-JJ_type_titre_id.pdf ; les documents du bail vont
    # dans un sous-dossier « bail_… », la pièce du locataire à la racine.
    aujourdhui = datetime.now(timezone.utc).date().isoformat()
    racine = [n for n in fichiers if "/" not in n]
    sous_bail = [n for n in fichiers if n.startswith("bail_")]
    assert len(racine) == 1 and len(sous_bail) == 2
    assert racine[0].startswith(f"{aujourdhui}_autre_Preuve_d'assurance")
    assert racine[0].endswith(".pdf")
    assert all(f"/{aujourdhui}_" in n for n in sous_bail)
    assert any("_bail_Bail_signé_2026_" in n for n in sous_bail)
    assert all(
        n.split("/")[0].endswith(f"_{seed['bail_id']}") for n in sous_bail
    )
    # Chaque fichier est le PDF d'origine.
    for n in fichiers:
        assert zf.read(n) == _PDF

    # index.csv : BOM + « ; », une ligne par fichier, colonnes attendues.
    lignes = _lignes_csv(zf.read("index.csv"))
    assert lignes[0] == [
        "Date", "Type", "Titre", "Source", "Envoyé le", "Envoyé à",
        "Ouvert le", "Signé le", "Signé par", "Fichier",
    ]
    assert len(lignes) - 1 == len(fichiers)
    assert {l[-1] for l in lignes[1:]} == set(fichiers)
    rappel = next(l for l in lignes[1:] if l[1] == "rappel_paiement")
    assert rappel[5] == "eliane@example.com" and rappel[3] == "genere"
    bail = next(l for l in lignes[1:] if l[1] == "bail")
    assert bail[8] == "Éliane Exporté" and bail[3] == "importe"


def test_zip_categorie_dossier_exclut_les_communications(
    client, auth_headers, seed
):
    r = client.get(
        f"/api/v1/immobilier/locataires/{seed['locataire_id']}"
        "/documents.zip?categorie=dossier",
        headers=auth_headers,
    )
    assert r.status_code == 200, r.text
    noms = zipfile.ZipFile(io.BytesIO(r.content)).namelist()
    assert not any("rappel_paiement" in n for n in noms)
    assert len([n for n in noms if n != "index.csv"]) == 2
    assert (
        client.get(
            f"/api/v1/immobilier/locataires/{seed['locataire_id']}"
            "/documents.zip?categorie=nimporte",
            headers=auth_headers,
        ).status_code
        == 400
    )


def test_zip_bail_logement_immeuble(client, auth_headers, seed):
    # Bail : à plat, sans les pièces propres au locataire.
    r = client.get(
        f"/api/v1/immobilier/baux/{seed['bail_id']}/documents.zip",
        headers=auth_headers,
    )
    assert r.status_code == 200, r.text
    noms = zipfile.ZipFile(io.BytesIO(r.content)).namelist()
    assert len(noms) == 3 and all("/" not in n for n in noms)

    # Logement : ses baux en sous-dossier.
    r = client.get(
        f"/api/v1/immobilier/logements/{seed['logement_id']}/documents.zip",
        headers=auth_headers,
    )
    assert r.status_code == 200, r.text
    noms = zipfile.ZipFile(io.BytesIO(r.content)).namelist()
    assert len([n for n in noms if n.startswith("bail_")]) == 2

    # Immeuble : logement_<n°>/bail_…/ + locataire_<nom>_<id>/.
    r = client.get(
        f"/api/v1/immobilier/immeubles/{seed['immeuble_id']}/documents.zip",
        headers=auth_headers,
    )
    assert r.status_code == 200, r.text
    noms = zipfile.ZipFile(io.BytesIO(r.content)).namelist()
    fichiers = [n for n in noms if n != "index.csv"]
    assert len(fichiers) == 3
    assert len([n for n in fichiers if n.startswith("logement_EXP-1/bail_")]) == 2
    assert any(
        n.startswith(f"locataire_Éliane_Exporté_{seed['locataire_id']}/")
        for n in fichiers
    )

    # Aucun document → 404 explicite (pas un zip vide).
    r = client.get(
        f"/api/v1/immobilier/immeubles/{seed['immeuble_ext_id']}/documents.zip",
        headers=auth_headers,
    )
    assert r.status_code == 404
    assert (
        client.get(
            "/api/v1/immobilier/locataires/999999/documents.zip",
            headers=auth_headers,
        ).status_code
        == 404
    )


def test_zip_413_au_dela_de_max_docs(client, auth_headers, seed, monkeypatch):
    monkeypatch.setattr(immobilier_exports, "MAX_DOCS", 1)
    r = client.get(
        f"/api/v1/immobilier/locataires/{seed['locataire_id']}/documents.zip",
        headers=auth_headers,
    )
    assert r.status_code == 413, r.text
    assert "Trop de documents" in r.json()["detail"]


def test_zip_413_au_dela_de_max_bytes(client, auth_headers, seed, monkeypatch):
    monkeypatch.setattr(immobilier_exports, "MAX_ZIP_BYTES", len(_PDF) + 1)
    r = client.get(
        f"/api/v1/immobilier/locataires/{seed['locataire_id']}/documents.zip",
        headers=auth_headers,
    )
    assert r.status_code == 413, r.text
    assert "trop volumineux" in r.json()["detail"]
